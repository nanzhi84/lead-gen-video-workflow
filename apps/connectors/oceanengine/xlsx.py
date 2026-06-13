"""Read an OceanEngine RPA-exported ``.xlsx`` into header-keyed dict rows.

openpyxl is the parser. It is imported lazily and optionally so the rest of the
connector (and the test suite) still loads when openpyxl is absent; callers get
an explicit :class:`XlsxUnsupportedError` instead of an ``ImportError`` crash.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:  # pragma: no cover - typing only
    pass


class XlsxUnsupportedError(RuntimeError):
    """Raised when XLSX parsing is requested but openpyxl is not installed."""


def openpyxl_available() -> bool:
    """Return whether the optional openpyxl dependency can be imported."""

    try:
        import openpyxl  # noqa: F401
    except Exception:
        return False
    return True


def _unique_headers(raw: list[str]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, header in enumerate(raw, start=1):
        normalized = (header or "").strip() or f"column_{index}"
        count = seen.get(normalized, 0) + 1
        seen[normalized] = count
        headers.append(normalized if count == 1 else f"{normalized}_{count}")
    return headers


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_first_sheet(xlsx_path: str | Path) -> list[dict[str, str]]:
    """Parse the first worksheet of ``xlsx_path`` into a list of dict rows.

    The first non-empty row is the header. Trailing all-empty rows are skipped.
    Raises :class:`XlsxUnsupportedError` if openpyxl is unavailable and
    :class:`FileNotFoundError` if the path does not exist.
    """

    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(path)
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover - exercised only without openpyxl
        raise XlsxUnsupportedError(
            "openpyxl is required to parse OceanEngine XLSX exports"
        ) from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows_iter = sheet.iter_rows(values_only=True)
        header_row: list[str] | None = None
        records: list[dict[str, str]] = []
        for raw_row in rows_iter:
            cells = [_cell_text(cell) for cell in raw_row]
            if not any(cells):
                continue
            if header_row is None:
                header_row = _unique_headers(cells)
                continue
            padded = cells + [""] * max(0, len(header_row) - len(cells))
            records.append({header_row[i]: padded[i] for i in range(len(header_row))})
        return records
    finally:
        workbook.close()
